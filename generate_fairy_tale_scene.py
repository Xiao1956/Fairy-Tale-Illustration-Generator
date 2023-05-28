import random
import os
import openpyxl


def load_characters(characters_path):
    """
    Load characters from a file.

    Args:
        characters_path (str): Path to the file containing fairytale characters.

    Returns:
        list: A list of characters.
    """
    with open(characters_path, 'r') as file:
        characters = [line.strip() for line in file]
    return characters


def load_locations(locations_path):
    """
    Load locations from an Excel file.

    Args:
        locations_path (str): Path to the Excel file containing locations.

    Returns:
        list: A list of location dictionaries, each containing 'Location', 'Adjectives', and 'Details'.
    """
    locations = []
    workbook = openpyxl.load_workbook(locations_path)
    worksheet = workbook.active
    column_names = ['Location', 'Adjectives', 'Details']
    column_indexes = {cell.value: index for index, cell in enumerate(worksheet[1], start=1) if cell.value in column_names}

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        row_data = {column_name: row[column_index - 1] for column_name, column_index in column_indexes.items()}
        locations.append(row_data)

    workbook.close()
    return locations


def load_interactions(interactions_path1, interactions_path2):
    """
    Load interactions from two sources: an Excel file and a text file.

    Args:
        interactions_path1 (str): Path to the first Excel file containing interactions.
        interactions_path2 (str): Path to the second text file containing interactions.

    Returns:
        list: A list of interactions.
    """
    interactions = []
    workbook = openpyxl.load_workbook(interactions_path1)
    worksheet = workbook.active
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row[6] is not None:
            interactions.extend(row[6].split(", "))
    workbook.close()

    with open(interactions_path2, 'r') as file:
        interactions.extend([line.strip() for line in file])

    return interactions


def generate_fairytale_scene(characters, interactions, locations):
    """
    Generate a fairytale scene by randomly selecting characters, an interaction, and a location.

    Args:
        characters (list): A list of characters.
        interactions (list): A list of interactions.
        locations (list): A list of location dictionaries.

    Returns:
        str: A sentence describing the fairytale scene.
    """
    character1, character2 = random.sample(characters, 2)
    interaction = random.choice(interactions)
    location = random.choice(locations)
    adjective = random.choice(location['Adjectives'].split(", "))
    detail = random.choice(location['Details'].split(", "))

    article = "an" if adjective[0].lower() in ['a', 'e', 'i', 'o', 'u'] else "a"

    sentence = f"{character1.capitalize()} is {interaction} {character2.capitalize()} in {location['Location'].lower()}, {article} {adjective} place filled with {detail}."
    return sentence


def main():

    # Get the path of the current script file
    path = os.path.realpath(__file__)

    # Set the relative paths of the target files
    characters_path = os.path.join(os.path.dirname(path), 'data/Classic fairy tale characters.txt')
    locations_path = os.path.join(os.path.dirname(path), 'data/Locations.xlsx')
    interactions_path1 = os.path.join(os.path.dirname(path), 'data/Veale\'s location listing.xlsx')
    interactions_path2 = os.path.join(os.path.dirname(path), 'data/Interactions.txt')

    characters = load_characters(characters_path)
    locations = load_locations(locations_path)
    interactions = load_interactions(interactions_path1, interactions_path2)

    scene = generate_fairytale_scene(characters, interactions, locations)
    print(scene)
    return scene


if __name__ == '__main__':
    main()
